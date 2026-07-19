import pandas as pd
from openpyxl import load_workbook

from app.schema import SHEETS
from app.workbook import WorkbookStore


def test_workbook_initialization_and_safe_update(tmp_path):
    path = tmp_path / "data" / "Rental_Portfolio.xlsx"
    store = WorkbookStore(path)
    assert path.exists()
    assert set(SHEETS).issubset(load_workbook(path, read_only=True).sheetnames)

    store.update(
        {"Properties": pd.DataFrame([{"property_id": "p1", "name": "Property One"}])}
    )
    assert store.read("Properties").iloc[0]["name"] == "Property One"
    assert list((path.parent / "backups").glob("*.xlsx"))
    assert not path.with_suffix(".xlsx.lock").exists()


def test_schema_upgrade_preserves_existing_rows(tmp_path):
    path = tmp_path / "Rental_Portfolio.xlsx"
    pd.DataFrame([{"property_id": "p1", "name": "Keep me"}]).to_excel(
        path, sheet_name="Properties", index=False
    )
    store = WorkbookStore(path)
    assert store.read("Properties").iloc[0]["name"] == "Keep me"
    assert "Statements" in load_workbook(path, read_only=True).sheetnames
