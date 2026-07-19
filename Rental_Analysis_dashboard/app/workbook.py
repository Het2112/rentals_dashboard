"""Safe, workbook-backed persistence for the local dashboard."""

from __future__ import annotations

import os
import shutil
import tempfile
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill

from .schema import FINANCIAL_CLASSIFICATIONS, SHEETS


class WorkbookLockedError(RuntimeError):
    pass


class WorkbookStore:
    def __init__(self, path: Path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.backup_dir = self.path.parent / "backups"
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self.lock_path = self.path.with_suffix(self.path.suffix + ".lock")
        self.ensure_exists()

    def ensure_exists(self) -> None:
        if self.path.exists():
            self._upgrade_schema()
            return
        wb = Workbook()
        wb.remove(wb.active)
        for name, columns in SHEETS.items():
            ws = wb.create_sheet(name)
            ws.append(columns)
            self._style_header(ws)
        self._populate_lists(wb["Lists"])
        self._atomic_save(wb, backup=False)

    @staticmethod
    def _style_header(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _upgrade_schema(self) -> None:
        wb = load_workbook(self.path)
        changed = False
        for name, columns in SHEETS.items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name)
                ws.append(columns)
                self._style_header(ws)
                changed = True
                continue
            ws = wb[name]
            existing = [cell.value for cell in ws[1]] if ws.max_row else []
            if not any(existing):
                ws.delete_rows(1, ws.max_row)
                ws.append(columns)
                self._style_header(ws)
                changed = True
            else:
                for column in columns:
                    if column not in existing:
                        ws.cell(1, ws.max_column + 1, column)
                        changed = True
                self._style_header(ws)
        if changed:
            self._atomic_save(wb)

    def _populate_lists(self, ws) -> None:
        for classification, description in FINANCIAL_CLASSIFICATIONS.items():
            ws.append(["financial_classification", classification, description])

    @contextmanager
    def _lock(self):
        try:
            descriptor = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(descriptor, str(os.getpid()).encode())
            os.close(descriptor)
        except FileExistsError as exc:
            raise WorkbookLockedError(
                f"Workbook update already in progress. Remove {self.lock_path} only if no app is running."
            ) from exc
        try:
            yield
        finally:
            self.lock_path.unlink(missing_ok=True)

    def _atomic_save(self, wb, backup: bool = True) -> None:
        with self._lock():
            if backup and self.path.exists():
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                shutil.copy2(
                    self.path, self.backup_dir / f"{self.path.stem}_{stamp}.xlsx"
                )
            handle, temp_name = tempfile.mkstemp(suffix=".xlsx", dir=self.path.parent)
            os.close(handle)
            temp_path = Path(temp_name)
            try:
                wb.save(temp_path)
                check = load_workbook(temp_path, read_only=True)
                check.close()
                os.replace(temp_path, self.path)
            finally:
                temp_path.unlink(missing_ok=True)

    def read(self, sheet: str) -> pd.DataFrame:
        if sheet not in SHEETS:
            raise KeyError(f"Unknown worksheet: {sheet}")
        try:
            df = pd.read_excel(self.path, sheet_name=sheet, dtype=object)
        except ValueError:
            return pd.DataFrame(columns=SHEETS[sheet])
        return df.dropna(how="all").reindex(columns=SHEETS[sheet])

    def update(self, replacements: dict[str, pd.DataFrame]) -> None:
        wb = load_workbook(self.path)
        for sheet, frame in replacements.items():
            if sheet not in SHEETS:
                raise KeyError(f"Unknown worksheet: {sheet}")
            if sheet in wb.sheetnames:
                del wb[sheet]
            ws = wb.create_sheet(sheet)
            columns = SHEETS[sheet]
            ws.append(columns)
            normalized = frame.reindex(columns=columns)
            for row in normalized.itertuples(index=False, name=None):
                ws.append([None if pd.isna(value) else value for value in row])
            self._style_header(ws)
            for column_cells in ws.columns:
                width = min(
                    42,
                    max(
                        11, max(len(str(cell.value or "")) for cell in column_cells) + 2
                    ),
                )
                ws.column_dimensions[column_cells[0].column_letter].width = width
        self._atomic_save(wb)

    def append(self, sheet: str, records: Iterable[dict]) -> None:
        incoming = pd.DataFrame(list(records))
        if incoming.empty:
            return
        self.update({sheet: pd.concat([self.read(sheet), incoming], ignore_index=True)})

    def refresh_excel_dashboards(self, metrics: pd.DataFrame) -> None:
        portfolio = []
        if not metrics.empty:
            portfolio = [
                {
                    "metric": "Operating Revenue",
                    "value": metrics["operating_revenue"].sum(),
                },
                {
                    "metric": "Operating Expenses",
                    "value": metrics["operating_expenses"].sum(),
                },
                {"metric": "NOI", "value": metrics["noi"].sum()},
                {"metric": "CapEx", "value": metrics["capex"].sum()},
                {
                    "metric": "Cash Flow After Debt",
                    "value": metrics["cash_flow_after_debt"].sum(),
                },
            ]
        prop_rows = []
        if not metrics.empty:
            grouped = (
                metrics.groupby(["property_id", "property_name"], dropna=False)[
                    [
                        "operating_revenue",
                        "operating_expenses",
                        "noi",
                        "cash_flow_after_debt",
                    ]
                ]
                .sum()
                .reset_index()
            )
            for row in grouped.to_dict("records"):
                for metric in [
                    "operating_revenue",
                    "operating_expenses",
                    "noi",
                    "cash_flow_after_debt",
                ]:
                    prop_rows.append(
                        {
                            "property_id": row["property_id"],
                            "property_name": row["property_name"],
                            "metric": metric,
                            "value": row[metric],
                        }
                    )
        self.update(
            {
                "Portfolio_Dashboard": pd.DataFrame(portfolio),
                "Property_Dashboard": pd.DataFrame(prop_rows),
            }
        )
        wb = load_workbook(self.path)
        if portfolio:
            ws = wb["Portfolio_Dashboard"]
            chart = BarChart()
            chart.title = "Portfolio Performance"
            chart.add_data(
                Reference(ws, min_col=2, min_row=1, max_row=ws.max_row),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            )
            ws.add_chart(chart, "D2")
        if not metrics.empty:
            ws = wb["Monthly_Metrics"]
            if ws.max_row > 1:
                chart = LineChart()
                chart.title = "Monthly NOI and Cash Flow"
                noi_col = SHEETS["Monthly_Metrics"].index("noi") + 1
                cash_col = SHEETS["Monthly_Metrics"].index("cash_flow_after_debt") + 1
                chart.add_data(
                    Reference(
                        ws,
                        min_col=noi_col,
                        max_col=cash_col,
                        min_row=1,
                        max_row=ws.max_row,
                    ),
                    titles_from_data=True,
                )
                chart.set_categories(
                    Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                )
                ws.add_chart(chart, "Z2")
        self._atomic_save(wb)
